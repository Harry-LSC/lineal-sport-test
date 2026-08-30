#!/usr/bin/env python3
"""
Lineal Sport Championship — Excel -> JSON builder

Reads:
    LSC_Update_System.xlsx

Writes:
    data/sports-data.json

The script deliberately recalculates championship state from the Baseline
and Updates sheets instead of relying on Excel formula cache values.
"""

from __future__ import annotations

import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK = Path("LSC_Update_System.xlsx")
OUTPUT = Path("data/sports-data.json")

VALID_STATUSES = {
    "Scheduled",
    "Awaiting result",
    "Completed",
    "Postponed",
    "Cancelled",
}

VALID_RESULTS = {
    "Holder win",
    "Challenger win",
    "Draw / tie",
    "No result",
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def iso_date(value):
    if value in ("", None):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raise ValueError(f"Expected an Excel date, got {value!r}")


def fail(message):
    print(f"ERROR: {message}", file=sys.stderr)
    raise SystemExit(1)


def read_baseline(ws):
    headers = [clean(c.value) for c in ws[1]]
    required = [
        "Championship",
        "Sport",
        "Competition",
        "Baseline Holder",
        "Baseline Since",
        "Current Defences",
        "Transfers",
        "Total Defences",
    ]
    missing = [h for h in required if h not in headers]
    if missing:
        fail(f"Baseline sheet is missing columns: {', '.join(missing)}")

    idx = {name: headers.index(name) + 1 for name in required}
    result = {}

    for r in range(2, ws.max_row + 1):
        championship = clean(ws.cell(r, idx["Championship"]).value)
        if not championship:
            continue

        if championship in result:
            fail(f"Duplicate championship in Baseline: {championship}")

        holder = clean(ws.cell(r, idx["Baseline Holder"]).value)
        since = ws.cell(r, idx["Baseline Since"]).value

        result[championship] = {
            "championship": championship,
            "sport": clean(ws.cell(r, idx["Sport"]).value),
            "competition": clean(ws.cell(r, idx["Competition"]).value),
            "current_holder": holder,
            "current_since": iso_date(since),
            "current_defences": int(ws.cell(r, idx["Current Defences"]).value or 0),
            "transfers": int(ws.cell(r, idx["Transfers"]).value or 0),
            "total_defences": int(ws.cell(r, idx["Total Defences"]).value or 0),
        }

    if not result:
        fail("Baseline sheet contains no championships.")

    return result


def read_updates(ws, known_championships):
    headers = [clean(c.value) for c in ws[1]]
    required = [
        "Championship",
        "Fixture Date",
        "Opponent",
        "Venue",
        "Status",
        "Result Type",
        "Result / Score",
        "Source URL",
        "Notes",
    ]
    missing = [h for h in required if h not in headers]
    if missing:
        fail(f"Updates sheet is missing columns: {', '.join(missing)}")

    idx = {name: headers.index(name) + 1 for name in required}
    events = []

    for r in range(2, ws.max_row + 1):
        championship = clean(ws.cell(r, idx["Championship"]).value)
        fixture_date_raw = ws.cell(r, idx["Fixture Date"]).value
        opponent = clean(ws.cell(r, idx["Opponent"]).value)
        status = clean(ws.cell(r, idx["Status"]).value)

        # Ignore genuinely blank rows.
        if not any([championship, fixture_date_raw, opponent, status]):
            continue

        if not championship:
            fail(f"Updates row {r}: Championship is blank.")
        if championship not in known_championships:
            fail(f"Updates row {r}: Unknown championship '{championship}'.")
        if not fixture_date_raw:
            fail(f"Updates row {r}: Fixture Date is blank.")
        if not opponent:
            fail(f"Updates row {r}: Opponent is blank.")
        if status not in VALID_STATUSES:
            fail(f"Updates row {r}: Invalid Status '{status}'.")

        result_type = clean(ws.cell(r, idx["Result Type"]).value)
        score = clean(ws.cell(r, idx["Result / Score"]).value)

        if status == "Completed":
            if result_type not in VALID_RESULTS:
                fail(
                    f"Updates row {r}: Completed event requires a valid Result Type."
                )
            if result_type != "No result" and not score:
                fail(
                    f"Updates row {r}: Completed event requires Result / Score."
                )
        elif result_type:
            fail(
                f"Updates row {r}: Result Type should be blank until Status is Completed."
            )

        event = {
            "row": r,
            "championship": championship,
            "fixture_date": iso_date(fixture_date_raw),
            "opponent": opponent,
            "venue": clean(ws.cell(r, idx["Venue"]).value),
            "status": status,
            "result_type": result_type or None,
            "result_score": str(score) if score != "" else None,
            "source_url": clean(ws.cell(r, idx["Source URL"]).value) or None,
            "notes": clean(ws.cell(r, idx["Notes"]).value) or None,
        }
        events.append(event)

    # Chronology is critical. We allow different championships to interleave,
    # but within each championship dates must never move backwards.
    last_date = {}
    for event in events:
        champ = event["championship"]
        dt = event["fixture_date"]
        if champ in last_date and dt < last_date[champ]:
            fail(
                f"Updates row {event['row']}: chronology goes backwards for "
                f"{champ} ({dt} < {last_date[champ]})."
            )
        last_date[champ] = dt

    return events


def build_state(baseline, events):
    states = {name: dict(values) for name, values in baseline.items()}
    lineage_updates = []
    pending = {name: [] for name in baseline}

    for event in events:
        champ = event["championship"]
        state = states[champ]
        status = event["status"]

        if status == "Completed":
            holder_before = state["current_holder"]
            result_type = event["result_type"]

            if result_type == "Challenger win":
                outcome = "Transfer"
                holder_after = event["opponent"]
                state["current_holder"] = holder_after
                state["current_since"] = event["fixture_date"]
                state["current_defences"] = 0
                state["transfers"] += 1

            elif result_type in {"Holder win", "Draw / tie"}:
                outcome = "Defence"
                holder_after = holder_before
                state["current_defences"] += 1
                state["total_defences"] += 1

            elif result_type == "No result":
                outcome = "No result"
                holder_after = holder_before

            else:
                fail(f"Internal error: unexpected result type {result_type!r}")

            lineage_updates.append(
                {
                    "championship": champ,
                    "date": event["fixture_date"],
                    "holder_before": holder_before,
                    "challenger": event["opponent"],
                    "result_type": result_type,
                    "result_score": event["result_score"],
                    "outcome": outcome,
                    "holder_after": holder_after,
                    "venue": event["venue"] or None,
                    "source_url": event["source_url"],
                    "notes": event["notes"],
                }
            )

        elif status in {"Scheduled", "Awaiting result"}:
            pending[champ].append(event)

        # Postponed and Cancelled events deliberately do not affect state.

    championships = []

    for champ, state in states.items():
        candidates = sorted(
            pending[champ],
            key=lambda e: (e["fixture_date"], e["row"]),
        )

        next_event = candidates[0] if candidates else None

        championships.append(
            {
                **state,
                "next_defence": (
                    {
                        "date": next_event["fixture_date"],
                        "opponent": next_event["opponent"],
                        "venue": next_event["venue"] or None,
                        "status": next_event["status"],
                    }
                    if next_event
                    else None
                ),
            }
        )

    return {
        "schema_version": 1,
        "championships": championships,
        "lineage_updates": lineage_updates,
    }


def main():
    if not WORKBOOK.exists():
        fail(f"Workbook not found: {WORKBOOK}")

    wb = load_workbook(WORKBOOK, data_only=False, read_only=True)

    for sheet in ("Baseline", "Updates"):
        if sheet not in wb.sheetnames:
            fail(f"Workbook is missing required sheet: {sheet}")

    baseline = read_baseline(wb["Baseline"])
    events = read_updates(wb["Updates"], baseline)
    output = build_state(baseline, events)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(
        json.dumps(output, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(
        f"Built {OUTPUT}: "
        f"{len(output['championships'])} championships, "
        f"{len(output['lineage_updates'])} completed update event(s)."
    )


if __name__ == "__main__":
    main()
